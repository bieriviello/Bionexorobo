"""
╔══════════════════════════════════════════════════════════╗
║           BIONEXO BOT — Interface de Automação           ║
║   Responde cotações automaticamente usando seu catálogo  ║
╚══════════════════════════════════════════════════════════╝

Dependências:
    pip install customtkinter openpyxl

Como usar:
    python bionexo_bot.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import customtkinter as ctk
import threading
import json
import os
import time
import re
import datetime
from bionexo_data import DataManager
from bionexo_engine import BionexoBotEngine
from pathlib import Path


try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Tema ──────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

# ── Paleta de cores ───────────────────────────────────────
COR_VERDE       = "#1F538D"
COR_VERDE_CLARO = "#E1EAF5"
COR_VERDE_MED   = "#5D9BCA"
COR_VERDE_ATIVO = "#14375E"
COR_CINZA_BG    = "#F7F6F3"
COR_CINZA_CARD  = "#FFFFFF"
COR_CINZA_BORDA = "#E0DED8"
COR_TEXTO       = "#2C2C2A"
COR_TEXTO_SEC   = "#6B6B68"
COR_AMARELO     = "#F59E0B"
COR_VERMELHO    = "#E24B4A"
COR_AZUL        = "#378ADD"

# ── Arquivo de configuração ───────────────────────────────
CONFIG_FILE = "bionexo_config.json"
CATALOGO_CACHE = "catalogo_cache.json"




# ══════════════════════════════════════════════════════════
#  JANELA PRINCIPAL
# ══════════════════════════════════════════════════════════

class BionexoApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("Bionexo Bot")
        self.geometry("1100x720")
        self.minsize(900, 600)
        self.configure(fg_color=COR_CINZA_BG)

        self.config = DataManager.carregar_config()
        callbacks = {
            "log": self._log,
            "metric_add": self._atualizar_metrica,
            "finish_cycle": self._bionexo_engine_finish
        }
        self.catalogo = []          # lista de dicts com os produtos
        self.engine = BionexoBotEngine(self.config, self.catalogo, callbacks)
        self.bot_rodando = False
        self.thread_bot = None
        self.logs = []

        self._build_ui()

        # Carrega catálogo do cache ou do arquivo original
        cache = DataManager.carregar_catalogo_cache()
        if cache is not None:
            self.catalogo = cache
            self.after(300, self._atualizar_tabela)
            self.after(300, self._atualizar_contador)
        elif self.config.get("arquivo_catalogo") and os.path.exists(self.config["arquivo_catalogo"]):
            self.after(300, lambda: self._carregar_catalogo_arquivo(self.config["arquivo_catalogo"], silencioso=True))

    # ── Layout principal ──────────────────────────────────

    def _build_ui(self):
        # Sidebar
        self.sidebar = ctk.CTkFrame(self, width=220, fg_color=COR_VERDE, corner_radius=0)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)
        self._build_sidebar()

        # Conteúdo
        self.frame_conteudo = ctk.CTkFrame(self, fg_color=COR_CINZA_BG, corner_radius=0)
        self.frame_conteudo.pack(side="left", fill="both", expand=True)

        # Páginas
        self.paginas = {}
        for nome in ("catalogo", "bot", "historico", "config"):
            frame = ctk.CTkFrame(self.frame_conteudo, fg_color=COR_CINZA_BG, corner_radius=0)
            frame.place(relx=0, rely=0, relwidth=1, relheight=1)
            self.paginas[nome] = frame

        self._build_pagina_catalogo()
        self._build_pagina_bot()
        self._build_pagina_historico()
        self._build_pagina_config()

        self._mostrar_pagina("catalogo")

    # ── Sidebar ───────────────────────────────────────────

    def _build_sidebar(self):
        # Logo / título
        logo = ctk.CTkLabel(
            self.sidebar,
            text="B  I  O  N  E  X  O",
            font=ctk.CTkFont(family="Courier", size=13, weight="bold"),
            text_color="#FFFFFF",
        )
        logo.pack(pady=(28, 2))
        sub = ctk.CTkLabel(
            self.sidebar, text="BOT DE COTAÇÕES",
            font=ctk.CTkFont(size=10),
            text_color="#9FE1CB",
        )
        sub.pack(pady=(0, 28))

        ctk.CTkFrame(self.sidebar, height=1, fg_color="#2D8A65").pack(fill="x", padx=20, pady=(0, 20))

        nav_items = [
            ("catalogo",  "📦  Catálogo"),
            ("bot",       "🤖  Robô"),
            ("historico", "📋  Histórico"),
            ("config",    "⚙️   Configurações"),
        ]
        self.nav_buttons = {}
        for key, label in nav_items:
            btn = ctk.CTkButton(
                self.sidebar, text=label, anchor="w",
                font=ctk.CTkFont(size=13),
                fg_color="transparent", hover_color="#0F6E56",
                text_color="#FFFFFF", corner_radius=8,
                height=42, width=190,
                command=lambda k=key: self._mostrar_pagina(k),
            )
            btn.pack(padx=14, pady=3)
            self.nav_buttons[key] = btn

        # Status do bot
        self.label_status_bot = ctk.CTkLabel(
            self.sidebar, text="● Bot pausado",
            font=ctk.CTkFont(size=11),
            text_color="#9FE1CB",
        )
        self.label_status_bot.pack(side="bottom", pady=20)

        self.label_catalogo_count = ctk.CTkLabel(
            self.sidebar, text="0 produtos no catálogo",
            font=ctk.CTkFont(size=10),
            text_color="#9FE1CB",
        )
        self.label_catalogo_count.pack(side="bottom", pady=(0, 4))

    def _mostrar_pagina(self, nome):
        for key, frame in self.paginas.items():
            frame.lower()
        self.paginas[nome].lift()
        for key, btn in self.nav_buttons.items():
            btn.configure(fg_color=COR_VERDE_ATIVO if key == nome else "transparent")

    # ══════════════════════════════════════════════════════
    #  PÁGINA 1 — CATÁLOGO
    # ══════════════════════════════════════════════════════

    def _build_pagina_catalogo(self):
        p = self.paginas["catalogo"]

        # Cabeçalho
        header = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0, height=70,
                               border_width=0)
        header.pack(fill="x")
        header.pack_propagate(False)

        ctk.CTkLabel(
            header, text="Catálogo de Materiais",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=COR_TEXTO,
        ).pack(side="left", padx=28, pady=18)

        btn_importar = ctk.CTkButton(
            header, text="+ Importar Excel / CSV",
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COR_VERDE, hover_color=COR_VERDE_ATIVO,
            height=36, corner_radius=8,
            command=self._importar_catalogo,
        )
        btn_importar.pack(side="right", padx=28, pady=16)

        # Barra de filtros
        filtro_frame = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0, height=56,
                                     border_width=0)
        filtro_frame.pack(fill="x")
        filtro_frame.pack_propagate(False)

        ctk.CTkFrame(filtro_frame, height=1, fg_color=COR_CINZA_BORDA).pack(fill="x", side="top")

        inner_filtro = ctk.CTkFrame(filtro_frame, fg_color="transparent")
        inner_filtro.pack(fill="x", padx=24, pady=10)

        ctk.CTkLabel(inner_filtro, text="Filtrar:", font=ctk.CTkFont(size=12),
                     text_color=COR_TEXTO_SEC).pack(side="left")

        self.entry_filtro = ctk.CTkEntry(
            inner_filtro, placeholder_text="Digite produto, código ou marca...",
            font=ctk.CTkFont(size=12), height=32, width=320,
            border_color=COR_CINZA_BORDA, fg_color=COR_CINZA_BG,
        )
        self.entry_filtro.pack(side="left", padx=(8, 16))
        self.entry_filtro.bind("<KeyRelease>", lambda e: self._atualizar_tabela())

        self.var_ativo = ctk.StringVar(value="Todos")
        seg = ctk.CTkSegmentedButton(
            inner_filtro,
            values=["Todos", "Ativos", "Inativos"],
            variable=self.var_ativo,
            font=ctk.CTkFont(size=11),
            command=lambda _: self._atualizar_tabela(),
            fg_color=COR_CINZA_BG,
            selected_color=COR_VERDE,
            selected_hover_color=COR_VERDE_ATIVO,
        )
        seg.pack(side="left")

        self.label_contagem = ctk.CTkLabel(
            inner_filtro, text="0 produtos",
            font=ctk.CTkFont(size=11), text_color=COR_TEXTO_SEC,
        )
        self.label_contagem.pack(side="right")

        # Tabela
        tabela_frame = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0)
        tabela_frame.pack(fill="both", expand=True, padx=0, pady=0)

        ctk.CTkFrame(tabela_frame, height=1, fg_color=COR_CINZA_BORDA).pack(fill="x")

        # Treeview estilizada
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Bionexo.Treeview",
            background=COR_CINZA_CARD,
            fieldbackground=COR_CINZA_CARD,
            foreground=COR_TEXTO,
            rowheight=32,
            font=("Segoe UI", 10),
            borderwidth=0,
        )
        style.configure("Bionexo.Treeview.Heading",
            background=COR_CINZA_BG,
            foreground=COR_TEXTO_SEC,
            font=("Segoe UI", 10, "bold"),
            borderwidth=0,
            relief="flat",
        )
        style.map("Bionexo.Treeview",
            background=[("selected", COR_VERDE_CLARO)],
            foreground=[("selected", COR_VERDE)],
        )
        style.layout("Bionexo.Treeview", [('Treeview.treearea', {'sticky': 'nswe'})])

        colunas = ("descricao", "codigo", "preco", "prazo", "marca", "unidade", "estoque", "ativo")
        self.tree = ttk.Treeview(
            tabela_frame, columns=colunas, show="headings",
            style="Bionexo.Treeview", selectmode="browse",
        )

        heads = {
            "descricao": ("Descrição do Produto", 320, "w"),
            "codigo":    ("Código",  90, "center"),
            "preco":     ("Preço (R$)", 100, "center"),
            "prazo":     ("Prazo (dias)", 90, "center"),
            "marca":     ("Marca", 120, "w"),
            "unidade":   ("Unidade", 70, "center"),
            "estoque":   ("Estoque", 80, "center"),
            "ativo":     ("Ativo?", 60, "center"),
        }
        for col, (label, width, anchor) in heads.items():
            self.tree.heading(col, text=label)
            self.tree.column(col, width=width, anchor=anchor, minwidth=40)

        scroll_y = ttk.Scrollbar(tabela_frame, orient="vertical", command=self.tree.yview)
        scroll_x = ttk.Scrollbar(tabela_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)

        self.tree.tag_configure("par",   background="#FAFAF8")
        self.tree.tag_configure("impar", background="#FFFFFF")
        self.tree.tag_configure("inativo", foreground="#AAAAAA")

        # Barra inferior com ações de linha
        barra_baixo = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0, height=48,
                                    border_width=0)
        barra_baixo.pack(fill="x", side="bottom")
        ctk.CTkFrame(barra_baixo, height=1, fg_color=COR_CINZA_BORDA).pack(fill="x", side="top")

        ctk.CTkButton(
            barra_baixo, text="+ Adicionar produto manualmente",
            font=ctk.CTkFont(size=11), fg_color="transparent",
            hover_color=COR_CINZA_BG, text_color=COR_VERDE,
            height=32, command=self._adicionar_produto,
        ).pack(side="left", padx=16, pady=8)

        ctk.CTkButton(
            barra_baixo, text="Remover selecionado",
            font=ctk.CTkFont(size=11), fg_color="transparent",
            hover_color=COR_CINZA_BG, text_color=COR_VERMELHO,
            height=32, command=self._remover_produto,
        ).pack(side="left", padx=4, pady=8)

        ctk.CTkButton(
            barra_baixo, text="Exportar catálogo",
            font=ctk.CTkFont(size=11), fg_color="transparent",
            hover_color=COR_CINZA_BG, text_color=COR_TEXTO_SEC,
            height=32, command=self._exportar_catalogo,
        ).pack(side="right", padx=16, pady=8)

        # Área de import vazia (mostrada quando sem catálogo)
        self.frame_empty = ctk.CTkFrame(tabela_frame, fg_color=COR_CINZA_CARD, corner_radius=0)
        self.frame_empty.place(relx=0, rely=0, relwidth=1, relheight=1)

        ctk.CTkLabel(
            self.frame_empty,
            text="Nenhum catálogo importado",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color=COR_TEXTO_SEC,
        ).pack(expand=True, pady=(120, 4))
        ctk.CTkLabel(
            self.frame_empty,
            text="Clique em '+ Importar Excel / CSV' para carregar seus produtos",
            font=ctk.CTkFont(size=12),
            text_color=COR_CINZA_BORDA,
        ).pack()
        ctk.CTkButton(
            self.frame_empty,
            text="Importar agora",
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COR_VERDE, hover_color=COR_VERDE_ATIVO,
            height=40, width=180, corner_radius=8,
            command=self._importar_catalogo,
        ).pack(pady=16)

    # ── Importação de catálogo ─────────────────────────────

    def _importar_catalogo(self):
        path = filedialog.askopenfilename(
            title="Selecionar catálogo de produtos",
            filetypes=[
                ("Planilhas", "*.xlsx *.xls *.csv"),
                ("Excel", "*.xlsx *.xls"),
                ("CSV", "*.csv"),
                ("Todos", "*.*"),
            ]
        )
        if not path:
            return
        self._carregar_catalogo_arquivo(path)

    def _carregar_catalogo_arquivo(self, path, silencioso=False):
        try:
            ext = Path(path).suffix.lower()

            if ext == ".csv":
                import csv
                produtos = []
                with open(path, "r", encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        produtos.append(self._normalizar_linha(row))
            elif ext in (".xlsx", ".xls"):
                if not HAS_OPENPYXL:
                    messagebox.showerror("Erro", "Instale openpyxl:\npip install openpyxl")
                    return
                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    messagebox.showerror("Erro", "Planilha vazia.")
                    return
                headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])]
                produtos = []
                for row in rows[1:]:
                    if all(v is None or str(v).strip() == "" for v in row):
                        continue
                    d = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
                    produtos.append(self._normalizar_linha(d))
            else:
                messagebox.showerror("Formato inválido", "Use .xlsx, .xls ou .csv")
                return

            if not produtos:
                if not silencioso:
                    messagebox.showwarning("Aviso", "Nenhum produto encontrado no arquivo.")
                return

            self.catalogo = produtos
            self.config["arquivo_catalogo"] = path
            DataManager.salvar_config(self.config)
            DataManager.salvar_catalogo(self.catalogo)
            self._atualizar_tabela()
            self._atualizar_contador()

            if not silencioso:
                messagebox.showinfo("Sucesso", f"{len(produtos)} produtos importados com sucesso!")

        except Exception as e:
            if not silencioso:
                messagebox.showerror("Erro ao importar", str(e))

    def _normalizar_linha(self, d):
        """Tenta mapear colunas de qualquer nome para o padrão interno."""
        def buscar(chaves_possiveis, default=""):
            for k in chaves_possiveis:
                for dk in d:
                    if dk and k.lower() in str(dk).lower():
                        v = d[dk]
                        return str(v).strip() if v is not None else default
            return default

        preco_str = buscar(["preco", "preço", "valor", "price", "custo"])
        try:
            preco = float(re.sub(r"[^\d.,]", "", preco_str).replace(",", ".")) if preco_str else 0.0
        except Exception:
            preco = 0.0

        estoque_str = buscar(["estoque", "qtd", "quantidade", "stock", "saldo"])
        try:
            estoque = float(re.sub(r"[^\d.]", "", estoque_str)) if estoque_str else 999
        except Exception:
            estoque = 999

        prazo_str = buscar(["prazo", "entrega", "lead", "dias"])
        try:
            prazo = int(float(re.sub(r"[^\d.]", "", prazo_str))) if prazo_str else int(self.config.get("prazo_padrao", 3))
        except Exception:
            prazo = 3

        ativo_str = buscar(["ativo", "status", "active", "habilitado"]).upper()
        ativo = "NÃO" if ativo_str in ("NÃO", "NAO", "0", "FALSE", "N", "INATIVO", "INACTIVO") else "SIM"

        return {
            "descricao": buscar(["descricao", "descrição", "produto", "item", "material", "nome", "name"]),
            "codigo":    buscar(["codigo", "código", "cod", "code", "ref", "referência", "referencia", "sku"]),
            "preco":     preco,
            "prazo":     prazo,
            "marca":     buscar(["marca", "fabricante", "brand", "manufacturer", "fornecedor"]),
            "unidade":   buscar(["unidade", "un", "unit", "medida"], "UN"),
            "estoque":   estoque,
            "ativo":     ativo,
        }

    def _atualizar_tabela(self):
        termo = self.entry_filtro.get().lower().strip() if hasattr(self, "entry_filtro") else ""
        filtro_ativo = self.var_ativo.get() if hasattr(self, "var_ativo") else "Todos"

        for item in self.tree.get_children():
            self.tree.delete(item)

        visíveis = []
        for prod in self.catalogo:
            if filtro_ativo == "Ativos" and prod["ativo"] != "SIM":
                continue
            if filtro_ativo == "Inativos" and prod["ativo"] == "SIM":
                continue
            if termo:
                haystack = " ".join([
                    str(prod.get("descricao", "")),
                    str(prod.get("codigo", "")),
                    str(prod.get("marca", "")),
                ]).lower()
                if termo not in haystack:
                    continue
            visíveis.append(prod)

        for i, prod in enumerate(visíveis):
            tag = "par" if i % 2 == 0 else "impar"
            if prod["ativo"] != "SIM":
                tag = "inativo"
            preco_fmt = f"R$ {prod['preco']:.2f}" if prod["preco"] else "-"
            self.tree.insert("", "end", values=(
                prod["descricao"],
                prod["codigo"] or "-",
                preco_fmt,
                prod["prazo"],
                prod["marca"] or "-",
                prod["unidade"],
                int(prod["estoque"]) if str(prod["estoque"]).replace(".0","").isdigit() else prod["estoque"],
                prod["ativo"],
            ), tags=(tag,))

        self.label_contagem.configure(text=f"{len(visíveis)} produto{'s' if len(visíveis) != 1 else ''}")

        # Mostra/esconde tela vazia
        if self.catalogo:
            self.frame_empty.lower()
        else:
            self.frame_empty.lift()

    def _atualizar_contador(self):
        n = len(self.catalogo)
        self.label_catalogo_count.configure(text=f"{n} produto{'s' if n != 1 else ''} no catálogo")

    def _adicionar_produto(self):
        JanelaAdicionarProduto(self, self.catalogo, self._atualizar_tabela, self._atualizar_contador)

    def _remover_produto(self):
        sel = self.tree.selection()
        if not sel:
            return
        item = self.tree.item(sel[0])
        desc = item["values"][0]
        if messagebox.askyesno("Confirmar", f"Remover '{desc}' do catálogo?"):
            self.catalogo = [p for p in self.catalogo if p["descricao"] != desc]
            DataManager.salvar_catalogo(self.catalogo)
            self._atualizar_tabela()
            self._atualizar_contador()

    def _exportar_catalogo(self):
        if not self.catalogo:
            messagebox.showwarning("Aviso", "Catálogo vazio.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            title="Salvar catálogo",
        )
        if not path:
            return
        try:
            if not HAS_OPENPYXL:
                raise ImportError("openpyxl não instalado")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Catálogo"
            headers = ["Descrição", "Código", "Preço (R$)", "Prazo (dias)", "Marca", "Unidade", "Estoque", "Ativo?"]
            ws.append(headers)
            for p in self.catalogo:
                ws.append([p["descricao"], p["codigo"], p["preco"], p["prazo"],
                            p["marca"], p["unidade"], p["estoque"], p["ativo"]])
            wb.save(path)
            messagebox.showinfo("Exportado", f"Catálogo salvo em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    # ══════════════════════════════════════════════════════
    #  PÁGINA 2 — ROBÔ
    # ══════════════════════════════════════════════════════

    def _build_pagina_bot(self):
        p = self.paginas["bot"]

        # Header
        header = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0, height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(
            header, text="Robô de Cotações",
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color=COR_TEXTO,
        ).pack(side="left", padx=28, pady=18)

        # Corpo em duas colunas
        corpo = ctk.CTkFrame(p, fg_color="transparent")
        corpo.pack(fill="both", expand=True, padx=24, pady=20)

        col_esq = ctk.CTkFrame(corpo, fg_color="transparent")
        col_esq.pack(side="left", fill="both", expand=True, padx=(0, 12))

        col_dir = ctk.CTkFrame(corpo, fg_color="transparent")
        col_dir.pack(side="right", fill="both", expand=True, padx=(12, 0))

        # ─── Card: Controle do robô ─────────────────────
        card_bot = ctk.CTkFrame(col_esq, fg_color=COR_CINZA_CARD,
                                 corner_radius=12, border_width=1, border_color=COR_CINZA_BORDA)
        card_bot.pack(fill="x", pady=(0, 16))

        ctk.CTkLabel(card_bot, text="Controle do robô",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=COR_TEXTO).pack(anchor="w", padx=20, pady=(18, 4))
        ctk.CTkLabel(card_bot, text="O robô faz login na Bionexo, busca cotações abertas e\nresponde automaticamente usando seu catálogo.",
                     font=ctk.CTkFont(size=11), text_color=COR_TEXTO_SEC,
                     justify="left").pack(anchor="w", padx=20, pady=(0, 16))

        # Indicador de status
        status_row = ctk.CTkFrame(card_bot, fg_color=COR_CINZA_BG, corner_radius=8)
        status_row.pack(fill="x", padx=20, pady=(0, 16))

        self.dot_status = ctk.CTkLabel(status_row, text="●", font=ctk.CTkFont(size=18),
                                        text_color="#AAAAAA")
        self.dot_status.pack(side="left", padx=(14, 6), pady=10)
        self.label_status_desc = ctk.CTkLabel(status_row, text="Robô parado",
                                               font=ctk.CTkFont(size=12, weight="bold"),
                                               text_color=COR_TEXTO_SEC)
        self.label_status_desc.pack(side="left", pady=10)
        self.label_prox_exec = ctk.CTkLabel(status_row, text="",
                                             font=ctk.CTkFont(size=10),
                                             text_color=COR_TEXTO_SEC)
        self.label_prox_exec.pack(side="right", padx=14, pady=10)

        # Botões
        btn_row = ctk.CTkFrame(card_bot, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(0, 20))

        self.btn_iniciar = ctk.CTkButton(
            btn_row, text="▶  Iniciar robô",
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COR_VERDE, hover_color=COR_VERDE_ATIVO,
            height=42, corner_radius=8,
            command=self._iniciar_bot,
        )
        self.btn_iniciar.pack(side="left", fill="x", expand=True, padx=(0, 8))

        self.btn_parar = ctk.CTkButton(
            btn_row, text="■  Parar",
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color="#E8E6E0", hover_color="#D5D3CC",
            text_color=COR_TEXTO, height=42, corner_radius=8,
            command=self._parar_bot, state="disabled",
        )
        self.btn_parar.pack(side="left", fill="x", expand=True, padx=(0, 0))

        # Botão rodar agora
        self.btn_agora = ctk.CTkButton(
            card_bot, text="⚡  Rodar uma vez agora (teste)",
            font=ctk.CTkFont(size=12),
            fg_color="transparent", hover_color=COR_CINZA_BG,
            text_color=COR_AZUL, height=36, corner_radius=8,
            border_width=1, border_color=COR_AZUL,
            command=self._rodar_agora,
        )
        self.btn_agora.pack(padx=20, pady=(0, 20), fill="x")

        # ─── Card: Log em tempo real ─────────────────────
        card_log = ctk.CTkFrame(col_esq, fg_color=COR_CINZA_CARD,
                                 corner_radius=12, border_width=1, border_color=COR_CINZA_BORDA)
        card_log.pack(fill="both", expand=True)

        log_header = ctk.CTkFrame(card_log, fg_color="transparent")
        log_header.pack(fill="x", padx=20, pady=(16, 8))
        ctk.CTkLabel(log_header, text="Log em tempo real",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=COR_TEXTO).pack(side="left")
        ctk.CTkButton(log_header, text="Limpar", font=ctk.CTkFont(size=10),
                      fg_color="transparent", hover_color=COR_CINZA_BG,
                      text_color=COR_TEXTO_SEC, height=24, width=60,
                      command=self._limpar_log).pack(side="right")

        self.txt_log = ctk.CTkTextbox(
            card_log, font=ctk.CTkFont(family="Courier", size=11),
            fg_color="#1A1A18", text_color="#5DCAA5",
            corner_radius=8, border_width=0,
        )
        self.txt_log.pack(fill="both", expand=True, padx=16, pady=(0, 16))
        self.txt_log.configure(state="disabled")
        self._log("Sistema iniciado. Configure suas credenciais e importe o catálogo.")

        # ─── Cards de métricas (coluna direita) ──────────
        metricas = [
            ("respondidas_hoje", "Respondidas hoje", "0", COR_VERDE),
            ("total_itens",      "Total de itens",   "0", COR_AZUL),
            ("sem_match",        "Sem match",        "0", COR_AMARELO),
            ("erros",            "Erros",            "0", COR_VERMELHO),
        ]
        self.labels_metricas = {}
        for i, (key, label, valor, cor) in enumerate(metricas):
            card = ctk.CTkFrame(col_dir, fg_color=COR_CINZA_CARD,
                                 corner_radius=12, border_width=1, border_color=COR_CINZA_BORDA)
            card.pack(fill="x", pady=(0, 12))
            ctk.CTkLabel(card, text=label, font=ctk.CTkFont(size=11),
                         text_color=COR_TEXTO_SEC).pack(anchor="w", padx=20, pady=(16, 2))
            lbl = ctk.CTkLabel(card, text=valor, font=ctk.CTkFont(size=32, weight="bold"),
                               text_color=cor)
            lbl.pack(anchor="w", padx=20, pady=(0, 16))
            self.labels_metricas[key] = lbl

        # Card ciclos
        card_ciclo = ctk.CTkFrame(col_dir, fg_color=COR_CINZA_CARD,
                                   corner_radius=12, border_width=1, border_color=COR_CINZA_BORDA)
        card_ciclo.pack(fill="x", pady=(0, 12))
        ctk.CTkLabel(card_ciclo, text="Intervalo de varredura",
                     font=ctk.CTkFont(size=11), text_color=COR_TEXTO_SEC).pack(anchor="w", padx=20, pady=(16, 4))
        self.label_intervalo_display = ctk.CTkLabel(
            card_ciclo,
            text=f"A cada {self.config.get('intervalo_min', '10')} minutos",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color=COR_TEXTO,
        )
        self.label_intervalo_display.pack(anchor="w", padx=20, pady=(0, 4))
        ctk.CTkLabel(card_ciclo, text="Altere nas Configurações",
                     font=ctk.CTkFont(size=10), text_color=COR_TEXTO_SEC).pack(anchor="w", padx=20, pady=(0, 16))

        # Métricas internas
        self._metricas = {"respondidas_hoje": 0, "total_itens": 0, "sem_match": 0, "erros": 0}

    def _log(self, msg, tipo="info"):
        hora = datetime.datetime.now().strftime("%H:%M:%S")
        prefixos = {"info": "  ", "ok": "✔ ", "erro": "✘ ", "aviso": "⚠ "}
        linha = f"[{hora}] {prefixos.get(tipo, '')} {msg}\n"
        self.logs.append(linha)
        self.txt_log.configure(state="normal")
        self.txt_log.insert("end", linha)
        self.txt_log.see("end")
        self.txt_log.configure(state="disabled")

    def _limpar_log(self):
        self.txt_log.configure(state="normal")
        self.txt_log.delete("1.0", "end")
        self.txt_log.configure(state="disabled")

    def _atualizar_metrica(self, key, delta=1):
        self._metricas[key] = self._metricas.get(key, 0) + delta
        self.labels_metricas[key].configure(text=str(self._metricas[key]))

    # ── Lógica do bot ─────────────────────────────────────


    def _bionexo_engine_finish(self):
        self.after(0, lambda: self._set_estado_bot(False))

    def _iniciar_bot(self):
        if not self._validar_pre_inicio():
            return
        self._set_estado_bot(True)
        self.thread_bot = threading.Thread(target=self.engine.iniciar, daemon=True)
        self.thread_bot.start()

    def _parar_bot(self):
        self.engine.parar()
        self._set_estado_bot(False)
        self._log("Robô parado pelo usuário.", "aviso")

    def _rodar_agora(self):
        if not self._validar_pre_inicio():
            return
        self._log("▶ Execução manual iniciada...")
        threading.Thread(target=self.engine.testar_uma_vez, daemon=True).start()

    def _build_pagina_historico(self):
        p = self.paginas["historico"]

        header = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0, height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="Histórico de Execuções",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     text_color=COR_TEXTO).pack(side="left", padx=28, pady=18)

        ctk.CTkButton(header, text="Atualizar",
                      font=ctk.CTkFont(size=12),
                      fg_color="transparent", hover_color=COR_CINZA_BG,
                      text_color=COR_VERDE, height=32,
                      command=self._recarregar_historico).pack(side="right", padx=28)

        ctk.CTkButton(header, text="Limpar histórico",
                      font=ctk.CTkFont(size=12),
                      fg_color="transparent", hover_color=COR_CINZA_BG,
                      text_color=COR_VERMELHO, height=32,
                      command=self._limpar_historico).pack(side="right", padx=4)

        tabela_frame = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0)
        tabela_frame.pack(fill="both", expand=True)

        ctk.CTkFrame(tabela_frame, height=1, fg_color=COR_CINZA_BORDA).pack(fill="x")

        style = ttk.Style()
        style.configure("Hist.Treeview",
            background=COR_CINZA_CARD, fieldbackground=COR_CINZA_CARD,
            foreground=COR_TEXTO, rowheight=30, font=("Segoe UI", 10), borderwidth=0,
        )
        style.configure("Hist.Treeview.Heading",
            background=COR_CINZA_BG, foreground=COR_TEXTO_SEC,
            font=("Segoe UI", 10, "bold"), borderwidth=0, relief="flat",
        )
        style.map("Hist.Treeview",
            background=[("selected", COR_VERDE_CLARO)],
            foreground=[("selected", COR_VERDE)],
        )

        colunas = ("data", "cotacoes", "respondidos", "sem_match", "status")
        self.tree_hist = ttk.Treeview(tabela_frame, columns=colunas,
                                       show="headings", style="Hist.Treeview")
        heads_hist = {
            "data":        ("Data / Hora",         180, "w"),
            "cotacoes":    ("Cotações analisadas",  150, "center"),
            "respondidos": ("Itens respondidos",    150, "center"),
            "sem_match":   ("Sem match",            120, "center"),
            "status":      ("Status",               120, "center"),
        }
        for col, (label, width, anchor) in heads_hist.items():
            self.tree_hist.heading(col, text=label)
            self.tree_hist.column(col, width=width, anchor=anchor)

        sc = ttk.Scrollbar(tabela_frame, orient="vertical", command=self.tree_hist.yview)
        self.tree_hist.configure(yscrollcommand=sc.set)
        sc.pack(side="right", fill="y")
        self.tree_hist.pack(fill="both", expand=True)

        self.tree_hist.tag_configure("ok",    foreground="#1D6B4A")
        self.tree_hist.tag_configure("erro",  foreground=COR_VERMELHO)
        self.tree_hist.tag_configure("vazio", foreground=COR_TEXTO_SEC)

        self._recarregar_historico()

    def _recarregar_historico(self):
        for item in self.tree_hist.get_children():
            self.tree_hist.delete(item)
        try:
            if not os.path.exists("bionexo_historico.json"):
                return
            # with open("bionexo_historico.json", "r", encoding="utf-8") as f:
                hist = DataManager.carregar_historico()
            for entrada in hist:
                tag = "ok" if entrada["status"] == "OK" else ("erro" if "Erro" in entrada["status"] else "vazio")
                self.tree_hist.insert("", "end", values=(
                    entrada["timestamp"],
                    entrada["cotacoes"],
                    entrada["respondidos"],
                    entrada["sem_match"],
                    entrada["status"],
                ), tags=(tag,))
        except Exception:
            pass

    def _limpar_historico(self):
        if messagebox.askyesno("Confirmar", "Limpar todo o histórico?"):
            if os.path.exists("bionexo_historico.json"):
                DataManager.limpar_historico()
            self._recarregar_historico()

    # ══════════════════════════════════════════════════════
    #  PÁGINA 4 — CONFIGURAÇÕES
    # ══════════════════════════════════════════════════════

    def _build_pagina_config(self):
        p = self.paginas["config"]

        header = ctk.CTkFrame(p, fg_color=COR_CINZA_CARD, corner_radius=0, height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        ctk.CTkLabel(header, text="Configurações",
                     font=ctk.CTkFont(size=20, weight="bold"),
                     text_color=COR_TEXTO).pack(side="left", padx=28, pady=18)

        scroll = ctk.CTkScrollableFrame(p, fg_color="transparent")
        scroll.pack(fill="both", expand=True, padx=24, pady=20)

        def card(titulo):
            f = ctk.CTkFrame(scroll, fg_color=COR_CINZA_CARD,
                              corner_radius=12, border_width=1, border_color=COR_CINZA_BORDA)
            f.pack(fill="x", pady=(0, 16))
            ctk.CTkLabel(f, text=titulo,
                         font=ctk.CTkFont(size=14, weight="bold"),
                         text_color=COR_TEXTO).pack(anchor="w", padx=20, pady=(18, 12))
            return f

        def campo(parent, label, placeholder, show="", inicial=""):
            row = ctk.CTkFrame(parent, fg_color="transparent")
            row.pack(fill="x", padx=20, pady=(0, 12))
            ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=12),
                         text_color=COR_TEXTO_SEC, width=160, anchor="w").pack(side="left")
            entry = ctk.CTkEntry(row, placeholder_text=placeholder,
                                  show=show, height=36, font=ctk.CTkFont(size=12),
                                  border_color=COR_CINZA_BORDA, fg_color=COR_CINZA_BG)
            entry.pack(side="left", fill="x", expand=True)
            if inicial:
                entry.insert(0, inicial)
            return entry

        # ─── Credenciais ─────────────────────────────────
        c1 = card("Credenciais Bionexo")
        self.e_email = campo(c1, "E-mail de acesso", "seu@email.com.br", inicial=self.config.get("email", ""))
        self.e_senha = campo(c1, "Senha", "••••••••", show="•", inicial=self.config.get("senha", ""))
        self.e_cnpj  = campo(c1, "CNPJ", "00.000.000/0001-00", inicial=self.config.get("cnpj", ""))

        ctk.CTkLabel(c1,
            text="  Suas credenciais são salvas localmente neste computador, nunca enviadas a terceiros.",
            font=ctk.CTkFont(size=10), text_color=COR_TEXTO_SEC,
        ).pack(anchor="w", padx=20, pady=(0, 16))

        # ─── Regras de resposta ───────────────────────────
        c2 = card("Regras de Resposta")
        self.e_margem = campo(c2, "Margem extra (%)", "0", inicial=self.config.get("margem", "0"))
        self.e_prazo  = campo(c2, "Prazo padrão (dias)", "3", inicial=self.config.get("prazo_padrao", "3"))
        self.e_intervalo = campo(c2, "Intervalo (minutos)", "10", inicial=self.config.get("intervalo_min", "10"))

        ctk.CTkLabel(c2,
            text="  Margem extra: ex. 5 = sobe 5% no preço antes de enviar. 0 = sem alteração.",
            font=ctk.CTkFont(size=10), text_color=COR_TEXTO_SEC,
        ).pack(anchor="w", padx=20, pady=(0, 16))

        # ─── Notificações ─────────────────────────────────
        c3 = card("Notificações por E-mail")
        row_notif = ctk.CTkFrame(c3, fg_color="transparent")
        row_notif.pack(fill="x", padx=20, pady=(0, 12))
        ctk.CTkLabel(row_notif, text="Notificar por e-mail",
                     font=ctk.CTkFont(size=12), text_color=COR_TEXTO_SEC,
                     width=160, anchor="w").pack(side="left")
        self.sw_notif = ctk.CTkSwitch(row_notif, text="",
                                       onvalue=True, offvalue=False,
                                       progress_color=COR_VERDE)
        self.sw_notif.pack(side="left")
        if self.config.get("notificar_email"):
            self.sw_notif.select()

        self.e_email_notif = campo(c3, "E-mail de destino", "seu@email.com.br",
                                    inicial=self.config.get("email_notificacao", ""))
        ctk.CTkFrame(c3, height=1).pack(pady=(0, 4))

        # ─── Botão salvar ─────────────────────────────────
        ctk.CTkButton(
            scroll, text="Salvar Configurações",
            font=ctk.CTkFont(size=14, weight="bold"),
            fg_color=COR_VERDE, hover_color=COR_VERDE_ATIVO,
            height=46, corner_radius=10,
            command=self._salvar_config_ui,
        ).pack(fill="x", pady=(8, 0))

    def _salvar_config_ui(self):
        self.config.update({
            "email":             self.e_email.get().strip(),
            "senha":             self.e_senha.get().strip(),
            "cnpj":              self.e_cnpj.get().strip(),
            "margem":            self.e_margem.get().strip() or "0",
            "prazo_padrao":      self.e_prazo.get().strip() or "3",
            "intervalo_min":     self.e_intervalo.get().strip() or "10",
            "notificar_email":   bool(self.sw_notif.get()),
            "email_notificacao": self.e_email_notif.get().strip(),
        })
        DataManager.salvar_config(self.config)
        self.label_intervalo_display.configure(
            text=f"A cada {self.config['intervalo_min']} minutos"
        )
        messagebox.showinfo("Salvo", "Configurações salvas com sucesso!")


# ══════════════════════════════════════════════════════════
#  JANELA — ADICIONAR PRODUTO MANUAL
# ══════════════════════════════════════════════════════════

class JanelaAdicionarProduto(ctk.CTkToplevel):
    def __init__(self, parent, catalogo, callback_tabela, callback_contador):
        super().__init__(parent)
        self.title("Adicionar Produto")
        self.geometry("480x650")
        self.resizable(False, True)
        self.configure(fg_color=COR_CINZA_BG)
        self.grab_set()

        self.catalogo = catalogo
        self.callback_tabela = callback_tabela
        self.callback_contador = callback_contador

        ctk.CTkLabel(self, text="Adicionar produto manualmente",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color=COR_TEXTO).pack(padx=24, pady=(20, 16), anchor="w")

        frame = ctk.CTkFrame(self, fg_color=COR_CINZA_CARD, corner_radius=12,
                              border_width=1, border_color=COR_CINZA_BORDA)
        frame.pack(fill="x", padx=20, pady=(0, 16))

        def campo(label, placeholder, row_frame=frame):
            row = ctk.CTkFrame(row_frame, fg_color="transparent")
            row.pack(fill="x", padx=16, pady=(10, 0))
            ctk.CTkLabel(row, text=label, font=ctk.CTkFont(size=11),
                         text_color=COR_TEXTO_SEC).pack(anchor="w")
            e = ctk.CTkEntry(row, placeholder_text=placeholder,
                              height=34, font=ctk.CTkFont(size=12),
                              border_color=COR_CINZA_BORDA, fg_color=COR_CINZA_BG)
            e.pack(fill="x", pady=(2, 0))
            return e

        self.e_desc   = campo("Descrição *", "Ex: Luva de procedimento M caixa 100un")
        self.e_cod    = campo("Código interno", "Ex: LUV-M-100")
        self.e_preco  = campo("Preço de venda (R$) *", "Ex: 38.90")
        self.e_prazo  = campo("Prazo de entrega (dias)", "Ex: 3")
        self.e_marca  = campo("Marca / Fabricante", "Ex: Supermax")
        self.e_estoque = campo("Estoque disponível", "Ex: 500")

        row_un = ctk.CTkFrame(frame, fg_color="transparent")
        row_un.pack(fill="x", padx=16, pady=(10, 14))
        ctk.CTkLabel(row_un, text="Unidade", font=ctk.CTkFont(size=11),
                     text_color=COR_TEXTO_SEC).pack(anchor="w")
        self.opt_unidade = ctk.CTkOptionMenu(
            row_un, values=["UN", "CX", "FR", "PCT", "RL", "AMP", "BL", "KG", "LT", "MT"],
            height=34, font=ctk.CTkFont(size=12),
            fg_color=COR_CINZA_BG, button_color=COR_VERDE,
        )
        self.opt_unidade.pack(fill="x", pady=(2, 0))

        ctk.CTkButton(
            self, text="Adicionar produto",
            font=ctk.CTkFont(size=13, weight="bold"),
            fg_color=COR_VERDE, hover_color=COR_VERDE_ATIVO,
            height=42, corner_radius=8,
            command=self._salvar,
        ).pack(fill="x", padx=20, pady=8)

    def _salvar(self):
        desc  = self.e_desc.get().strip()
        preco_str = self.e_preco.get().strip().replace(",", ".")
        if not desc:
            messagebox.showwarning("Campo obrigatório", "Preencha a descrição.")
            return
        try:
            preco = float(preco_str)
        except ValueError:
            messagebox.showwarning("Valor inválido", "Preço inválido.")
            return

        self.catalogo.append({
            "descricao": desc,
            "codigo":    self.e_cod.get().strip(),
            "preco":     preco,
            "prazo":     int(self.e_prazo.get().strip() or "3"),
            "marca":     self.e_marca.get().strip(),
            "unidade":   self.opt_unidade.get(),
            "estoque":   float(self.e_estoque.get().strip() or "999"),
            "ativo":     "SIM",
        })
        DataManager.salvar_catalogo(self.catalogo)
        self.callback_tabela()
        self.callback_contador()
        self.destroy()


# ══════════════════════════════════════════════════════════
#  ENTRY POINT
# ══════════════════════════════════════════════════════════

if __name__ == "__main__":
    # Verifica dependências
    faltando = []
    try:
        import customtkinter
    except ImportError:
        faltando.append("customtkinter")
    try:
        import openpyxl
    except ImportError:
        faltando.append("openpyxl")

    if faltando:
        print("\n╔══════════════════════════════════════╗")
        print("║   BIONEXO BOT — Dependências         ║")
        print("╠══════════════════════════════════════╣")
        print("║  Execute o comando abaixo e tente    ║")
        print("║  iniciar novamente:                  ║")
        print("║                                      ║")
        print(f"║  pip install {' '.join(faltando)}")
        print("╚══════════════════════════════════════╝\n")
    else:
        app = BionexoApp()
        app.mainloop()
